from django.views.generic import TemplateView
from django.http import JsonResponse
from django.views import View
from app.models import Invitados,SugerenciaCancion
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import json

class IndexView(TemplateView):
    template_name = 'index.html'

class ConfirmarAsistencia(View):
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        if 'telefono' in data:
            code = data.get('telefono')
            es_invitado = Invitados.objects.filter(code=code)
            if es_invitado.exists():
                invitados_list = [
                    {'id': invitado.id, 'nombre': invitado.nombre_completo, 'confirmado': invitado.confirmado}
                    for invitado in es_invitado
                ]
                return JsonResponse({'status': 'success', 'invitados': invitados_list})
            else:
                return JsonResponse({'status': 'error', 'message': 'No es invitado'}, status=404)
        
        elif 'invitados' in data:
            invitados_ids = data.get('invitados', [])
            Invitados.objects.filter(id__in=invitados_ids).update(confirmado=True)
            return JsonResponse({'status': 'success', 'message': 'Asistencia confirmada'})

        return JsonResponse({'status': 'error', 'message': 'Solicitud incorrecta'}, status=400)

    def get(self, request, *args, **kwargs):
        return JsonResponse({'status': 'ready'}, status=200)



class SugerirCancionView(View):
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        nombre_usuario = data.get('nombre_usuario')
        descripcion_sugerencia = data.get('descripcion_sugerencia')  
        link = data.get('link')
        sugerencia = SugerenciaCancion(
            nombre_usuario=nombre_usuario,
            nombre_cancion_autor=descripcion_sugerencia,
            link=link
        )
        sugerencia.save()

        return JsonResponse({'status': 'success', 'message': '¡Gracias por tu sugerencia!'})

class ExportSugerenciasExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sugerencias"
        ws.merge_cells('A1:C1')
        cell = ws['A1']
        cell.value = "Sugerencia de canciones"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        headers = ['Nombre del Usuario', 'Nombre de la Canción y Autor', 'Link']
        ws.append(headers)

        for cell in ws[2]:  
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        sugerencias = SugerenciaCancion.objects.all()
        for sugerencia in sugerencias:
            ws.append([sugerencia.nombre_usuario, sugerencia.nombre_cancion_autor, sugerencia.link])

        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column) 
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=sugerencias_canciones.xlsx"
        wb.save(response)
        return response
    
class ExportInvitadosExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invitados"

        ws.merge_cells('A1:C1')
        cell = ws['A1']
        cell.value = "Reporte de Invitados"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        headers = ['Nombre Completo',  'Confirmado']
        ws.append(headers)

        for cell in ws[2]:  
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        invitados = Invitados.objects.all()
        for invitado in invitados:
            ws.append([invitado.nombre_completo,"Sí" if invitado.confirmado else "No"])

        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=invitados.xlsx"
        wb.save(response)
        return response